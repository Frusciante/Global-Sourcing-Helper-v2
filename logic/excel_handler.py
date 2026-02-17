import os
import pandas as pd
import openpyxl
from rapidfuzz import process, fuzz
from collections import defaultdict

class ExcelHandler:
    def __init__(self, target_file, log_callback, config):
        self.target_file = target_file
        self.log_callback = log_callback
        self.config = config
        
        # 중복된 소분류를 모두 담기 위해 리스트를 값으로 갖는 사전 사용
        # { '소분류': ['경로1', '경로2', ...] }
        self.cp_map = defaultdict(list)
        self.nv_map = defaultdict(list)
        
        self.cp_leaf_nodes = []
        self.nv_leaf_nodes = []
        
        self.load_categories()

    def load_categories(self):
        try:
            if not os.path.exists(self.target_file): 
                self.log_callback(f"⚠️ [Excel] 파일 없음: {self.target_file}")
                return
            
            self.log_callback("📂 [Excel] 카테고리 중복 방지 지도 구축 중...")
            
            cp_df = pd.read_excel(self.target_file, sheet_name='쿠팡 전체 카테고리 (240517)', dtype=str)
            nv_df = pd.read_excel(self.target_file, sheet_name='네이버 전체 카테고리 (251215)', dtype=str)
            
            target_col = '여기서 카테고리를 복사해주세요'

            def build_map_list(df, col_name):
                mapping = defaultdict(list)
                col_data = df[col_name].dropna().tolist() if col_name in df.columns else df.iloc[:, 0].dropna().tolist()
                for path in col_data:
                    full_path = str(path).strip()
                    leaf = full_path.split('>')[-1].strip()
                    # 덮어쓰지 않고 리스트에 추가 (중복 경로 보존)
                    if full_path not in mapping[leaf]:
                        mapping[leaf].append(full_path)
                return mapping

            self.cp_map = build_map_list(cp_df, target_col)
            self.nv_map = build_map_list(nv_df, target_col)
            
            self.cp_leaf_nodes = list(self.cp_map.keys())
            self.nv_leaf_nodes = list(self.nv_map.keys())

            self.log_callback(f"✅ [Excel] 구축 완료 (항목: 쿠팡 {len(self.cp_leaf_nodes)}, 네이버 {len(self.nv_leaf_nodes)})")
        except Exception as e:
            self.log_callback(f"❌ [Excel] 로드 실패: {e}")

    def get_category_candidates(self, core_item, alt_item, full_title, shop_type='naver', limit=10):
        path_map = self.cp_map if shop_type == 'coupang' else self.nv_map
        leaf_nodes = self.cp_leaf_nodes if shop_type == 'coupang' else self.nv_leaf_nodes
        
        if not leaf_nodes: return []
    
        final_candidates = []
        seen_paths = set()
        forbidden_roots = ["도서", "잡지", "국내도서", "외국도서", "eBook", "중고", "만화"]
    
        # --- [1단계 & 2단계: 인간의 Ctrl+F 검색 및 정렬] ---
        search_targets = [t for t in [core_item, alt_item] if t]
        filtered_leaves = []
        
        for target in search_targets:
            # 소분류 명칭에 검색어가 포함된 것들을 모두 수집
            matched = [leaf for leaf in leaf_nodes if target in leaf]
            filtered_leaves.extend(matched)
    
        filtered_leaves = list(set(filtered_leaves)) # 중복 제거
    
        if filtered_leaves:
            # Ctrl+F로 찾은 리스트 안에서 상품명과 가장 어울리는 것 추출
            results = process.extract(full_title, filtered_leaves, scorer=fuzz.WRatio, limit=limit)
            for res in results:
                matched_leaf = res[0]
                paths = path_map.get(matched_leaf, [])
                for path in paths:
                    if any(root in path for root in forbidden_roots): continue
                    if path not in seen_paths:
                        final_candidates.append(path)
                        seen_paths.add(path)
    
        # --- [3단계: 최후의 수단 - 기존 명사 가산점 알고리즘 (Fallback)] ---
        # 1, 2단계에서 단어 포함 매칭이 단 하나도 안 되었을 때만 실행됩니다.
        if not final_candidates:
            self.log_callback(f"⚠️ '{core_item}'/'{alt_item}' 포함 단어 없음. 기존 가산점 알고리즘으로 전환합니다.")
            
            # 검색할 키워드 리스트 (상위 2개)
            keyword_list = search_targets
            quota_per_word = limit // max(1, len(keyword_list))
    
            for keyword in keyword_list:
                if len(final_candidates) >= limit: break
                
                # 30개 정도를 유사도 기반으로 우선 추출
                results = process.extract(keyword, leaf_nodes, scorer=fuzz.WRatio, limit=30)
                
                core_noun = keyword[-1] if keyword else ""
                scored_candidates = []
                
                for res in results:
                    matched_leaf, base_score = res[0], res[1]
                    adjusted_score = base_score
                    
                    # 가산점 부여 (끝 글자 일치 시 +50, 포함 시 +10)
                    if core_noun and matched_leaf.endswith(core_noun):
                        adjusted_score += 50 
                    elif core_noun and core_noun in matched_leaf:
                        adjusted_score += 10
                    
                    paths = path_map.get(matched_leaf, [])
                    for path in paths:
                        if any(root in path for root in forbidden_roots): continue
                        scored_candidates.append((path, adjusted_score))
                
                # 점수순 정렬 후 추가
                scored_candidates.sort(key=lambda x: x[1], reverse=True)
                added_count = 0
                for path, score in scored_candidates:
                    if path not in seen_paths:
                        final_candidates.append(path)
                        seen_paths.add(path)
                        added_count += 1
                    if added_count >= quota_per_word or len(final_candidates) >= limit:
                        break
        
        return final_candidates[:limit]


    def save_product(self, data_row):
        try:
            wb = openpyxl.load_workbook(self.target_file)
            ws = wb['엑셀 수집 양식 (Ver.9)']
            
            start_row = 7
            while ws.cell(row=start_row, column=4).value is not None: start_row += 1
            
            tags_value = data_row.get('tags', '')
            if isinstance(tags_value, list): tags_value = ", ".join(tags_value)
            
            ws.cell(row=start_row, column=2, value=data_row.get('cp_cat', ''))
            ws.cell(row=start_row, column=3, value=data_row.get('nv_cat', ''))
            ws.cell(row=start_row, column=4, value=data_row.get('translated_title', ''))
            ws.cell(row=start_row, column=5, value=tags_value)
            ws.cell(row=start_row, column=6, value=data_row.get('url', ''))
            
            try:
                cost_basic = int(self.config.get('COST_BASIC', 3000))
                cost_exchange = int(self.config.get('COST_EXCHANGE', 6000))
                cost_return = int(self.config.get('COST_RETURN', 6000))
            except: cost_basic, cost_exchange, cost_return = 3000, 6000, 6000

            ws.cell(row=start_row, column=7, value=0)
            ws.cell(row=start_row, column=8, value='유료' if cost_basic > 0 else '무료')
            ws.cell(row=start_row, column=9, value=cost_basic)
            ws.cell(row=start_row, column=10, value=cost_exchange)
            ws.cell(row=start_row, column=11, value=cost_return)
            
            ws.cell(row=start_row, column=12, value=data_row.get('manufacturer', 'OEM'))
            ws.cell(row=start_row, column=13, value=data_row.get('brand', 'OEM'))
            ws.cell(row=start_row, column=14, value=data_row.get('model', ''))
            
            wb.save(self.target_file)
            self.log_callback(f"💾 [Excel] 저장 완료 (행: {start_row})")
            return True
            
        except PermissionError:
            self.log_callback("❌ [Excel] 저장 실패: 엑셀 파일을 닫아주세요.")
            return False
        except Exception as e:
            self.log_callback(f"❌ [Excel] 오류: {e}")
            return False